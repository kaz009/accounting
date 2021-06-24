# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd

#CSV→Excel
data = pd.read_csv('rakuten.csv')
data.to_excel('rakuten.xlsx', encoding='utf-8')
data = pd.read_csv('tap.csv')
data.to_excel('tap.xlsx', encoding='utf-8')


wb_rakuten = openpyxl.load_workbook("rakuten.xlsx", data_only=True)
ws_rakuten = wb_rakuten.worksheets[0]
wb_tap = openpyxl.load_workbook("tap.xlsx", data_only=True)
ws_tap = wb_tap.worksheets[0]
wb_result = openpyxl.Workbook()
ws_result = wb_result.active

#A列削除、最終行取得
ws_rakuten.delete_cols(1)
ws_tap.delete_cols(1)
rakuten_row = ws_rakuten.max_row
tap_row = ws_tap.max_row



#エクセルへ書き込み
resultRow = 1
#楽天データ取り込み
rakuten={}
tap={}

for i in range(2,rakuten_row+1):
    rakuten_no=ws_rakuten["M"+str(i)].value
    rakuten_no=rakuten_no.strip("'")
    rakuten_price=ws_rakuten["U"+str(i)].value
    if rakuten_no is None or rakuten_price is None:
        continue    
    elif rakuten_no in rakuten.keys():
        old_price=rakuten[rakuten_no]
        new_price=int(old_price)+int(rakuten_price)
        rakuten[rakuten_no]=new_price
    else:
        rakuten[rakuten_no]=rakuten_price
        
        
        
#タップデータ取り込み
for i in range(3,tap_row+1):
    tap_no=ws_tap["N"+str(i)].value
    if tap_no is not None:
        if " " in tap_no:
            tap_no=tap_no.split(" ")[1]
    tap_price=ws_tap["J"+str(i)].value
    tap_price=str(tap_price)
    tap_price=tap_price.replace(",","")
    tap_price=int(tap_price)
    if tap_no is None or tap_price is None:
        pass   
    elif tap_no in tap.keys():
        old_price=tap[tap_no]
        new_price=int(old_price)+int(tap_price)
        tap[tap_no]=new_price
    else:
        tap[tap_no]=tap_price

print(tap)
print(rakuten)
#楽天→タップ
rakuten_len=len(rakuten)
tap_len=len(tap)
for i in rakuten:
    if i in tap:
        continue
    else:
        ws_result.cell(column= 1, row=resultRow,value=i+":"+str(rakuten[i])+"(タップ側にない)")
        resultRow += 1
         
        
        
#タップ→楽天
for i in tap:
    if i in rakuten:
        if tap[i]==rakuten[i]:
            continue
        else:
            ws_result.cell(column=1, row=resultRow, value=i+":"+str(rakuten[i])+"(金額が違います)")
            resultRow += 1
        
    else:
        ws_result.cell(column=1, row=resultRow, value=i+":"+str(tap[i])+"(楽天側にない)")
        resultRow += 1
wb_result.save('RakutenResult.xlsx')
            
#外部予約番号がない分に関してはわからない        
      
        
        
        
  