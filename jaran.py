# -*- coding: utf-8 -*-"""

import pandas as pd
import openpyxl

#CSV→Excel
data = pd.read_csv('jaran.csv',encoding='CP932')
data.to_excel('jaran.xlsx', encoding='utf-8')
data = pd.read_csv('tap.CSV', encoding='CP932')
data.to_excel('tap.xlsx', encoding='utf-8')


wb_jaran = openpyxl.load_workbook("jaran.xlsx", data_only=True)
ws_jaran = wb_jaran.worksheets[0]
wb_tap = openpyxl.load_workbook("tap.xlsx", data_only=True)
ws_tap = wb_tap.worksheets[0]
wb_result = openpyxl.Workbook()
ws_result = wb_result.active


#A列削除、最終行取得
ws_jaran.delete_cols(1)
ws_tap.delete_cols(1) 
jaran_row = ws_jaran.max_row
tap_row = ws_tap.max_row


#じゃらんデータ取り込み
jaran={}
tap={}
jaran_price=0
for i in range(2,jaran_row+1):
    jaran_no=ws_jaran["D"+str(i)].value
    price1=ws_jaran["N"+str(i)].value
    price2=ws_jaran["O"+str(i)].value
    price3=ws_jaran["Q"+str(i)].value
    price4=ws_jaran["S"+str(i)].value
    price5=ws_jaran["T"+str(i)].value
    price6=ws_jaran["U"+str(i)].value
    if jaran_no is not None:
        if price1 is not None:
            price1=str(price1)
            price1=price1.replace("P","")
            price1=price1.replace(",","")
            jaran_price+=int(price1)
        if price2 is not None:
            price2=str(price2)
            price2=price2.replace(",","")
            jaran_price+=int(price2)
        if price3 is not None:
            price3=str(price3)
            price3=price3.replace(",","")
            jaran_price+=int(price3)
        if price4 is not None:
            price4=str(price4)
            price4=price4.replace(",","")
            jaran_price+=int(price4)
        if price5 is not None:
            price5=str(price5)
            price5=price5.replace(",","")
            jaran_price+=int(price5)
        if price6 is not None:
            price6=str(price6)
            price6=price6.replace(",","")
            jaran_price+=int(price6)

        if jaran_no in jaran.keys():
            old_price=jaran[jaran_no]
            new_price=int(old_price)+int(jaran_price)
            jaran[jaran_no]=new_price
        else:
            jaran[jaran_no]=jaran_price
        jaran_price=0

        

#タップデータ取り込み
for i in range(3,tap_row+1):
    tap_no=ws_tap["N"+str(i)].value
    tap_price=ws_tap["J"+str(i)].value
    tap_price=str(tap_price)
    tap_price=tap_price.replace(",","")
    tap_price=int(tap_price)
    if tap_no is None or tap_price is None:
        continue    
    elif tap_no in tap.keys():
        old_price=tap[tap_no]
        new_price=int(old_price)+int(tap_price)
        tap[tap_no]=new_price
    else:
        tap[tap_no]=tap_price

#エクセルへ書き込み
resultRow = 1
#じゃらん→タップ
jaran_len=len(jaran)
tap_len=len(tap)
for i in jaran:
    if i in tap:
        continue
    else:
        ws_result.cell(column=1, row=resultRow, value=str(i)+" : "+str(jaran[i])+"円(タップ側にない)")
        resultRow += 1
        
        
#タップ→じゃらん
for i in tap:
    if i in jaran:
        if tap[i]==jaran[i]:
            continue
        else:
            ws_result.cell(column=1, row=resultRow, value=str(i)+" : "+str(jaran[i])+"円(金額が違います)")
        resultRow += 1
    else:
        ws_result.cell(column=1, row=resultRow, value=str(i)+" : "+str(tap[i])+"円(じゃらん側にない)")
        resultRow += 1

wb_result.save('JaranResult.xlsx')
            
#外部予約番号がない分に関しては計算できない為事前に入れ込むこと        
      
        
        